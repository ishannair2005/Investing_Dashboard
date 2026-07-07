"""
excel_workbook.py

The persistence layer: reads/writes a local .xlsx workbook (via
openpyxl) that is the actual "dashboard" the spec describes. This
replaces an earlier Google Sheets-based design -- a local file needs no
cloud service account, sharing step, or API quota, and opens directly
in Excel/Numbers.

Every write function here follows one of two patterns, chosen per the
spec's own rules:

  - APPEND-ONLY (Financials, Ratios, News, Narrative): existing rows are
    read first, new data is filtered down to only rows not already
    present, and only those are appended. Historical quarters/
    headlines/writeups are never overwritten or deleted.

  - UPSERT (Valuation, Dashboard, Watchlist's auto-columns): these tabs
    represent current-state snapshots, not history, so a ticker's row
    is found and overwritten in place (or appended if it doesn't exist
    yet). Watchlist's manual columns (thesis, catalysts, target price,
    etc.) are the one exception within an upsert tab -- ensure_watchlist_row
    never touches them once a row exists, since those are edited by
    hand in Excel, not by this program.

Unlike a live API, the workbook is loaded into memory once per process
and only written to disk explicitly (save_workbook()), which
sync_ticker_full() calls after each ticker so a crash partway through a
24-ticker run doesn't lose progress on the tickers already processed.

This module owns "what does the workbook currently contain" (needed to
decide what's new) and "how is each tab shaped" (headers). It calls
into financials.py/ratios.py/news.py's fetch functions to get data, but
receives already-generated AI analysis (analysis.py) as a parameter,
since that's expensive to (re)compute and orchestration of when to run
it belongs to main.py/add_company.py.
"""

import datetime as dt
import logging
import time
from typing import Optional

import openpyxl
import pandas as pd
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

import analysis
import financials
import news
import ratios
import tickers
import valuation
from analysis import NARRATIVE_FIELDS
from config import EXCEL_FILE_PATH
from utils import today_str

logger = logging.getLogger(__name__)

# --------------------------------------------------------------------------
# Tab schemas -- each header list is the single source of truth for that
# tab's column order. Financials/Ratios/Valuation reuse the field-order
# constants their owning modules expose, so this file can't silently
# drift out of sync with what those modules actually compute.
# --------------------------------------------------------------------------
FINANCIALS_TAB = "Financials"
FINANCIALS_HEADERS = ["ticker", "quarter", "period_end"] + financials.ALL_METRICS

RATIOS_TAB = "Ratios"
RATIOS_HEADERS = ["ticker", "quarter", "period_end"] + ratios.RATIO_METRICS

VALUATION_TAB = "Valuation"
VALUATION_HEADERS = ["ticker", "as_of"] + valuation.VALUATION_METRICS

NEWS_TAB = "News"
NEWS_HEADERS = news.NEWS_COLUMNS

NARRATIVE_TAB = "Narrative"
NARRATIVE_HEADERS = ["ticker", "generated_at"] + NARRATIVE_FIELDS

DASHBOARD_TAB = "Dashboard"
DASHBOARD_HEADERS = [
    "ticker", "name", "sector", "industry", "description",
    "price", "market_cap", "pe_ratio", "forward_pe", "peg_ratio", "dividend_yield",
    "revenue_growth_yoy", "net_margin", "roe", "debt_to_equity",
    "valuation_rating", "latest_recommendation", "last_updated",
]

WATCHLIST_TAB = "Watchlist"
# First 3 columns are auto-maintained (kept in sync with tickers.py); the
# remaining 6 are manual and this module never writes to them once a
# ticker's row exists.
WATCHLIST_AUTO_HEADERS = ["ticker", "company_name", "sector"]
WATCHLIST_MANUAL_HEADERS = ["investment_thesis", "catalysts", "risks", "target_price", "personal_rating", "notes"]
WATCHLIST_HEADERS = WATCHLIST_AUTO_HEADERS + WATCHLIST_MANUAL_HEADERS

# Tab order matters here -- it's the order sheets appear in the workbook,
# so the most-glanced-at tabs (Dashboard, Watchlist) come first.
ALL_TABS = {
    DASHBOARD_TAB: DASHBOARD_HEADERS,
    WATCHLIST_TAB: WATCHLIST_HEADERS,
    VALUATION_TAB: VALUATION_HEADERS,
    RATIOS_TAB: RATIOS_HEADERS,
    FINANCIALS_TAB: FINANCIALS_HEADERS,
    NEWS_TAB: NEWS_HEADERS,
    NARRATIVE_TAB: NARRATIVE_HEADERS,
}

# --------------------------------------------------------------------------
# Presentation -- colors, number formats, and column categories shared by
# every sheet, so "make it look neat" is one set of rules applied
# uniformly rather than per-tab special-casing.
# --------------------------------------------------------------------------
_HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")  # slate/charcoal
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)

# Column categories -> Excel number format. Membership-based (not a
# per-header dict) because the same field name (e.g. "revenue_growth_yoy")
# recurs across Ratios and Dashboard and should always format the same way.
_CURRENCY_PER_SHARE = {"price", "week_52_high", "week_52_low", "eps", "diluted_eps", "target_price"}
_CURRENCY_LARGE = {
    "market_cap", "enterprise_value", "revenue", "gross_profit", "operating_income", "ebit", "ebitda",
    "net_income", "cash", "total_assets", "current_assets", "total_debt", "long_term_debt",
    "shareholder_equity", "current_liabilities", "inventory", "operating_cash_flow",
    "capital_expenditures", "free_cash_flow", "investing_cash_flow", "financing_cash_flow", "interest_expense",
}
_PERCENTAGE = {
    "gross_margin", "operating_margin", "net_margin", "fcf_margin",
    "revenue_growth_qoq", "revenue_growth_yoy", "eps_growth_yoy",
    "roe", "roa", "roic", "dividend_yield",
}
_PLAIN_NUMBER = {
    "pe_ratio", "forward_pe", "peg_ratio", "price_to_sales", "price_to_book",
    "ev_to_revenue", "ev_to_ebitda", "debt_to_equity", "current_ratio", "quick_ratio",
    "interest_coverage", "beta",
}
_INTEGER_COMMAS = {"shares_outstanding", "average_volume", "employees"}
_SCORE_1DP = {"confidence_score", "competitive_position_score", "long_term_outlook_score"}
_DATE_FIELDS = {"as_of", "period_end", "generated_at", "last_updated", "date", "added_date"}

# Columns that hold long free-text (news summaries, AI writeup sections) --
# capped at a readable width instead of ballooning to fit a whole paragraph.
_LONG_TEXT_FIELDS = {
    "headline", "ai_summary", "description", "executive_summary", "key_developments",
    "bull_case", "bear_case", "financial_health", "valuation_assessment", "competitive_position",
    "long_term_outlook", "investment_thesis", "confidence_rationale", "latest_recommendation",
    "investment_thesis", "catalysts", "risks", "notes", "link",
}

_TABLE_STYLE = TableStyleInfo(
    name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False,
    showFirstColumn=False, showLastColumn=False,
)


def _number_format_for(header: str) -> Optional[str]:
    if header in _CURRENCY_PER_SHARE:
        return '$#,##0.00;[Red]-$#,##0.00'
    if header in _CURRENCY_LARGE:
        return '$#,##0;[Red]-$#,##0'
    if header in _PERCENTAGE:
        return '0.0%;[Red]-0.0%'
    if header in _PLAIN_NUMBER:
        return '#,##0.00;[Red]-#,##0.00'
    if header in _INTEGER_COMMAS:
        return '#,##0'
    if header in _SCORE_1DP:
        return '0.0'
    if header in _DATE_FIELDS:
        return 'yyyy-mm-dd'
    return None


def format_value_for_display(header: str, value) -> str:
    """Human-readable string for `value` under `header`, using the same
    column-category rules as the Excel number formats (_number_format_for)
    -- the single source of truth both this module's .xlsx output and
    dashboard_app.py's on-screen tables draw from, so the two never
    disagree on how a figure is presented.
    """
    if value is None:
        return "—"
    try:
        if pd.isna(value):
            return "—"
    except (TypeError, ValueError):
        pass

    if header in _CURRENCY_PER_SHARE:
        return f"${value:,.2f}"
    if header in _CURRENCY_LARGE:
        return f"${value:,.0f}"
    if header in _PERCENTAGE:
        return f"{value * 100:,.1f}%"
    if header in _PLAIN_NUMBER:
        return f"{value:,.2f}"
    if header in _INTEGER_COMMAS:
        return f"{value:,.0f}"
    if header in _SCORE_1DP:
        return f"{value:,.1f}"
    if header in _DATE_FIELDS:
        try:
            return pd.Timestamp(value).strftime("%Y-%m-%d")
        except (TypeError, ValueError):
            return str(value)
    return str(value)


def dehumanize_columns(df: pd.DataFrame, tab_name: str) -> pd.DataFrame:
    """Rename a DataFrame's columns from the humanized display labels
    written to row 1 (e.g. "P/E Ratio") back to the canonical snake_case
    field names (e.g. "pe_ratio") every function in this module keys
    off of.

    Needed by any external reader of the workbook (dashboard_app.py uses
    pandas.read_excel()) -- pandas takes column names straight from row
    1's literal text, which is the humanized label, not the internal
    field name. This is the read-side counterpart to _humanize_header().
    """
    headers = ALL_TABS.get(tab_name)
    if headers is None:
        return df
    reverse_map = {_humanize_header(h): h for h in headers}
    return df.rename(columns=reverse_map)


def _parse_date_maybe(value):
    """Convert a "YYYY-MM-DD" string into a real date object so Excel
    treats it as a date (sortable/filterable) instead of text. Falls
    back to the original string if it doesn't parse cleanly."""
    if not isinstance(value, str):
        return value
    try:
        return dt.datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError:
        return value


_workbook: Optional[openpyxl.Workbook] = None


# --------------------------------------------------------------------------
# Workbook setup
# --------------------------------------------------------------------------

def _get_workbook() -> openpyxl.Workbook:
    """Load the workbook from disk if it exists, else start a fresh one
    in memory. Unlike the Google Sheets version of this module, a local
    file can simply be created -- there's no "orphaned in a service
    account's Drive" problem to avoid.
    """
    global _workbook
    if _workbook is not None:
        return _workbook

    if EXCEL_FILE_PATH.exists():
        _workbook = openpyxl.load_workbook(EXCEL_FILE_PATH)
        logger.info("Loaded existing workbook: %s", EXCEL_FILE_PATH)
    else:
        _workbook = openpyxl.Workbook()
        logger.info("Starting new workbook: %s", EXCEL_FILE_PATH)
    return _workbook


def reload_workbook() -> None:
    """Drop the cached in-memory workbook so the next access re-reads it
    from disk. This module loads the workbook once per process and
    holds it in memory (_workbook) for the rest of that process's life
    -- correct for a short-lived CLI run, but a long-running Streamlit
    session can outlive a `git pull` that changed the file on disk out
    from under it. git_sync.sync_before_write() calls this right after
    pulling, so a write action starts from what's actually on GitHub
    rather than a stale in-process copy.
    """
    global _workbook
    _workbook = None


def _get_or_create_worksheet(tab_name: str, headers: list) -> Worksheet:
    wb = _get_workbook()

    if tab_name in wb.sheetnames:
        ws = wb[tab_name]
        if ws.max_row < 1 or ws.cell(row=1, column=1).value is None:
            _write_header(ws, headers)
        return ws

    # openpyxl.Workbook() starts with one default sheet named "Sheet" --
    # repurpose it for the first tab we create instead of leaving an
    # empty stray sheet in the final file.
    if wb.sheetnames == ["Sheet"] and wb["Sheet"].max_row <= 1 and wb["Sheet"].cell(1, 1).value is None:
        ws = wb["Sheet"]
        ws.title = tab_name
    else:
        ws = wb.create_sheet(title=tab_name)

    _write_header(ws, headers)
    logger.info("Created worksheet '%s' with %d columns", tab_name, len(headers))
    return ws


_ABBREVIATION_FIXUPS = {
    "Pe": "P/E", "Peg": "PEG", "Roe": "ROE", "Roa": "ROA", "Roic": "ROIC",
    "Eps": "EPS", "Ebit": "EBIT", "Ebitda": "EBITDA", "Yoy": "YoY", "Qoq": "QoQ",
    "Ev": "EV", "Fcf": "FCF", "Ai": "AI",
}


def _humanize_header(header: str) -> str:
    """"eps_growth_yoy" -> "EPS Growth YoY" style display label. Data
    underneath still keys off the raw snake_case name (see
    _get_all_records) -- this only affects the label shown in row 1.
    Word-level fixups rather than a per-field override dict, so a new
    field automatically gets sane casing without a matching entry here.
    """
    words = header.replace("_", " ").title().split(" ")
    return " ".join(_ABBREVIATION_FIXUPS.get(w, w) for w in words)


def _write_header(ws: Worksheet, headers: list) -> None:
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=_humanize_header(header))
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 26


def initialize_workbook() -> None:
    """Ensure every tab from ALL_TABS exists with its header row.
    Idempotent -- safe to call at the start of every run. Does not save
    to disk by itself; sync_ticker_full()/save_workbook() do that."""
    for tab_name, headers in ALL_TABS.items():
        _get_or_create_worksheet(tab_name, headers)
    logger.info("Workbook initialized: %s", ", ".join(ALL_TABS))


# --------------------------------------------------------------------------
# Presentation pass -- applied to every sheet right before every save, so
# the file is always fully formatted, not just after some "finalize" step
# that might get skipped.
# --------------------------------------------------------------------------

def _autosize_columns(ws: Worksheet, headers: list) -> None:
    """Size each column to its widest value, capped so long free-text
    columns (news headlines, AI writeup paragraphs) don't balloon the
    sheet width."""
    for col_idx, header in enumerate(headers, start=1):
        letter = get_column_letter(col_idx)
        header_display = ws.cell(row=1, column=col_idx).value
        max_len = len(str(header_display)) if header_display else 10
        for (cell,) in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        cap = 60 if header in _LONG_TEXT_FIELDS else 32
        ws.column_dimensions[letter].width = min(max_len + 2, cap)


def _apply_table(ws: Worksheet, tab_name: str, headers: list) -> None:
    """Turn the sheet's data range into a real Excel Table: banded rows
    plus built-in filter/sort dropdowns on every column -- structured,
    filterable data instead of a plain grid, which is most of what
    "Power BI vibes" actually means for a spreadsheet.

    Needs at least one data row -- a header-only table can corrupt the
    file in some Excel/openpyxl version combinations -- so this is
    skipped until real data exists, and re-applied (by widening .ref)
    on every save as more rows get appended.
    """
    if ws.max_row < 2:
        return

    ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    table_name = f"{tab_name.replace(' ', '')}Table"
    if table_name in ws.tables:
        ws.tables[table_name].ref = ref
    else:
        table = Table(displayName=table_name, ref=ref)
        table.tableStyleInfo = _TABLE_STYLE
        ws.add_table(table)


def _style_dashboard(ws: Worksheet) -> None:
    """Dashboard-specific polish: colored valuation-rating cells -- the
    at-a-glance visual element that makes this feel like a dashboard
    rather than a data export."""
    if ws.max_row < 2:
        return

    rating_col = DASHBOARD_HEADERS.index("valuation_rating") + 1
    rating_letter = get_column_letter(rating_col)
    last_row = ws.max_row

    rating_colors = {"Undervalued": "C6EFCE", "Overvalued": "FFC7CE", "Fairly Valued": "FFEB9C"}
    for rating, color in rating_colors.items():
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        ws.conditional_formatting.add(
            f"{rating_letter}2:{rating_letter}{last_row}",
            CellIsRule(operator="equal", formula=[f'"{rating}"'], fill=fill),
        )


def _apply_all_styling() -> None:
    wb = _get_workbook()
    for tab_name, headers in ALL_TABS.items():
        if tab_name not in wb.sheetnames:
            continue
        ws = wb[tab_name]
        _autosize_columns(ws, headers)
        _apply_table(ws, tab_name, headers)

    if DASHBOARD_TAB in wb.sheetnames:
        _style_dashboard(wb[DASHBOARD_TAB])
        wb.active = wb.sheetnames.index(DASHBOARD_TAB)


def save_workbook(retries: int = 3, backoff_seconds: float = 1.0) -> None:
    """Apply presentation styling, then persist the in-memory workbook
    to disk.

    Retries briefly on PermissionError -- the most common real-world
    cause is the file being open in Excel, and macOS/Windows can both
    momentarily lock a file during autosave/antivirus scans. If it's
    still locked after retrying, fail with an actionable message rather
    than a raw OS error.
    """
    wb = _get_workbook()
    _apply_all_styling()
    last_exc = None
    for attempt in range(1, retries + 1):
        try:
            wb.save(EXCEL_FILE_PATH)
            logger.info("Saved workbook to %s", EXCEL_FILE_PATH)
            return
        except PermissionError as exc:
            last_exc = exc
            if attempt < retries:
                time.sleep(backoff_seconds)
    raise RuntimeError(
        f"Could not save {EXCEL_FILE_PATH} -- it may be open in Excel. "
        f"Close the file and try again."
    ) from last_exc


# --------------------------------------------------------------------------
# Shared helpers
# --------------------------------------------------------------------------

def _to_cell(value, header: str):
    """Normalize a raw value for writing into a cell under `header`:
    pandas/NumPy NaN -> None (a genuinely blank cell -- Excel treats this
    correctly as empty in both text and numeric columns, unlike an empty
    string, which would make a numeric column display as text), and
    recognized date-field strings -> real date objects (so Excel sorts/
    filters them as dates instead of text).
    """
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    if header in _DATE_FIELDS:
        return _parse_date_maybe(value)
    return value


def _row_from_dict(row: dict, headers: list) -> list:
    return [_to_cell(row.get(h), h) for h in headers]


def _get_all_records(ws: Worksheet, headers: list) -> list:
    """Read every data row (below the header) into a list of dicts keyed
    by the canonical internal `headers` list -- the openpyxl equivalent
    of gspread's get_all_records().

    Takes `headers` explicitly rather than reading row 1's literal text:
    row 1 shows a humanized display label ("P/E Ratio"), which must stay
    decoupled from the snake_case internal field names ("pe_ratio") that
    every lookup in this module keys off of.
    """
    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        records.append(dict(zip(headers, row)))
    return records


def _apply_row_number_formats(ws: Worksheet, row_number: int, headers: list) -> None:
    for col_idx, header in enumerate(headers, start=1):
        fmt = _number_format_for(header)
        if fmt:
            ws.cell(row=row_number, column=col_idx).number_format = fmt


def _append_row(ws: Worksheet, row: list, headers: list) -> None:
    ws.append(row)
    _apply_row_number_formats(ws, ws.max_row, headers)


def _write_row_at(ws: Worksheet, row_number: int, row: list, headers: list) -> None:
    for col_idx, value in enumerate(row, start=1):
        ws.cell(row=row_number, column=col_idx, value=value)
    _apply_row_number_formats(ws, row_number, headers)


def _existing_quarters(ticker: str) -> set:
    ws = _get_or_create_worksheet(FINANCIALS_TAB, FINANCIALS_HEADERS)
    return {r["quarter"] for r in _get_all_records(ws, FINANCIALS_HEADERS) if r.get("ticker") == ticker}


def _existing_ratio_quarters(ticker: str) -> set:
    ws = _get_or_create_worksheet(RATIOS_TAB, RATIOS_HEADERS)
    return {r["quarter"] for r in _get_all_records(ws, RATIOS_HEADERS) if r.get("ticker") == ticker}


def _existing_news_links(ticker: str) -> set:
    ws = _get_or_create_worksheet(NEWS_TAB, NEWS_HEADERS)
    return {r["link"] for r in _get_all_records(ws, NEWS_HEADERS) if r.get("ticker") == ticker and r.get("link")}


def _find_row_number(ws: Worksheet, ticker: str) -> Optional[int]:
    """1-based sheet row number for `ticker`'s row, or None if it
    doesn't have one yet."""
    for row_idx in range(2, ws.max_row + 1):
        if ws.cell(row=row_idx, column=1).value == ticker:
            return row_idx
    return None


# --------------------------------------------------------------------------
# Financials (append-only)
# --------------------------------------------------------------------------

def sync_financials(ticker: str, financials_df: Optional[pd.DataFrame] = None) -> int:
    """Append any quarters for `ticker` not already in the Financials tab.
    Returns the number of rows appended.

    Accepts a pre-fetched financials_df (sync_ticker_full's shared
    fetch) to avoid a redundant yfinance call.
    """
    ticker = ticker.upper()
    ws = _get_or_create_worksheet(FINANCIALS_TAB, FINANCIALS_HEADERS)
    existing = _existing_quarters(ticker)

    new_quarters_df = financials.get_new_quarters(ticker, existing, df=financials_df)
    if new_quarters_df.empty:
        logger.info("Financials: no new quarters for %s", ticker)
        return 0

    for record in new_quarters_df.to_dict(orient="records"):
        _append_row(ws, _row_from_dict(record, FINANCIALS_HEADERS), FINANCIALS_HEADERS)
    logger.info("Financials: appended %d new quarter(s) for %s", len(new_quarters_df), ticker)
    return len(new_quarters_df)


# --------------------------------------------------------------------------
# Ratios (append-only)
# --------------------------------------------------------------------------

def sync_ratios(ticker: str, financials_df: Optional[pd.DataFrame] = None) -> int:
    """Compute the full ratio history for `ticker` (TTM ratios need the
    full quarterly series) and append only the quarters not already in
    the Ratios tab.

    Accepts a pre-fetched financials_df to avoid a redundant yfinance
    call when sync_financials() already fetched it this run.
    """
    ticker = ticker.upper()
    ws = _get_or_create_worksheet(RATIOS_TAB, RATIOS_HEADERS)
    existing = _existing_ratio_quarters(ticker)

    fin_df = financials_df if financials_df is not None else financials.get_quarterly_financials(ticker)
    ratios_df = ratios.compute_ratios(fin_df)
    if ratios_df.empty:
        logger.info("Ratios: no data for %s", ticker)
        return 0

    new_ratios_df = ratios_df[~ratios_df["quarter"].isin(existing)]
    if new_ratios_df.empty:
        logger.info("Ratios: no new quarters for %s", ticker)
        return 0

    for record in new_ratios_df.to_dict(orient="records"):
        _append_row(ws, _row_from_dict(record, RATIOS_HEADERS), RATIOS_HEADERS)
    logger.info("Ratios: appended %d new quarter(s) for %s", len(new_ratios_df), ticker)
    return len(new_ratios_df)


# --------------------------------------------------------------------------
# Valuation (upsert -- current snapshot, not history)
# --------------------------------------------------------------------------

def upsert_valuation(ticker: str, snapshot: Optional[dict] = None) -> None:
    ticker = ticker.upper()
    ws = _get_or_create_worksheet(VALUATION_TAB, VALUATION_HEADERS)
    snapshot = snapshot if snapshot is not None else valuation.get_valuation_snapshot(ticker)
    if not snapshot:
        logger.warning("Valuation: no data for %s -- skipping upsert", ticker)
        return

    row = _row_from_dict(snapshot, VALUATION_HEADERS)
    row_number = _find_row_number(ws, ticker)
    if row_number:
        _write_row_at(ws, row_number, row, VALUATION_HEADERS)
    else:
        _append_row(ws, row, VALUATION_HEADERS)
    logger.info("Valuation: upserted %s", ticker)


# --------------------------------------------------------------------------
# News (append-only, deduped by link)
# --------------------------------------------------------------------------

def sync_news(ticker: str, news_df: Optional[pd.DataFrame] = None) -> int:
    """Accepts a pre-fetched, already-classified news_df to avoid
    re-running (paid) AI classification on the same headlines twice in
    one run -- see sync_ticker_full."""
    ticker = ticker.upper()
    ws = _get_or_create_worksheet(NEWS_TAB, NEWS_HEADERS)
    existing_links = _existing_news_links(ticker)

    if news_df is None:
        new_news_df = news.get_new_news(ticker, existing_links)
    elif news_df.empty:
        new_news_df = news_df
    else:
        new_news_df = news_df[~news_df["link"].isin(existing_links)].reset_index(drop=True)

    if new_news_df.empty:
        logger.info("News: no new items for %s", ticker)
        return 0

    for record in new_news_df.to_dict(orient="records"):
        _append_row(ws, _row_from_dict(record, NEWS_HEADERS), NEWS_HEADERS)
    logger.info("News: appended %d new item(s) for %s", len(new_news_df), ticker)
    return len(new_news_df)


# --------------------------------------------------------------------------
# Narrative (append-only -- every generation is a dated history entry)
# --------------------------------------------------------------------------

def get_latest_narrative(ticker: str) -> Optional[dict]:
    """Most recent stored analysis for `ticker`, or None. Passed to
    analysis.generate_analysis() as `previous_analysis` so the model can
    identify what materially changed since last time."""
    ticker = ticker.upper()
    ws = _get_or_create_worksheet(NARRATIVE_TAB, NARRATIVE_HEADERS)
    records = [r for r in _get_all_records(ws, NARRATIVE_HEADERS) if r.get("ticker") == ticker]
    if not records:
        return None
    # generated_at is written as a real date object (see _DATE_FIELDS), so
    # the "missing" sentinel must be a date too -- comparing a date against
    # a string ("" ) would raise TypeError inside max().
    return max(records, key=lambda r: r.get("generated_at") or dt.date.min)


def append_narrative(ticker: str, ai_analysis: dict) -> None:
    """Always appends, never overwrites -- the Narrative tab is meant to
    preserve the full history of writeups over time, not just the
    latest one."""
    ticker = ticker.upper()
    ws = _get_or_create_worksheet(NARRATIVE_TAB, NARRATIVE_HEADERS)
    _append_row(ws, _row_from_dict(ai_analysis, NARRATIVE_HEADERS), NARRATIVE_HEADERS)
    logger.info("Narrative: appended new writeup for %s", ticker)


# --------------------------------------------------------------------------
# Dashboard (upsert -- rebuilt from the latest of everything else)
# --------------------------------------------------------------------------

def upsert_dashboard(
    ticker: str,
    profile: dict,
    valuation_snapshot: dict,
    latest_ratios: dict,
    latest_analysis: Optional[dict],
) -> None:
    """Rebuild one ticker's Dashboard row from already-fetched pieces.
    Takes the raw pieces rather than fetching them itself so callers
    (main.py) that already pulled this data for other tabs don't pay
    for redundant network/AI calls."""
    ticker = ticker.upper()
    ws = _get_or_create_worksheet(DASHBOARD_TAB, DASHBOARD_HEADERS)

    recommendation = None
    valuation_rating = None
    if latest_analysis:
        valuation_rating = latest_analysis.get("valuation_rating")
        thesis = latest_analysis.get("investment_thesis") or ""
        recommendation = (thesis[:200] + "...") if len(thesis) > 200 else thesis

    row_data = {
        "ticker": ticker,
        "name": (profile or {}).get("name"),
        "sector": (profile or {}).get("sector"),
        "industry": (profile or {}).get("industry"),
        "description": (profile or {}).get("description"),
        "price": (valuation_snapshot or {}).get("price"),
        "market_cap": (valuation_snapshot or {}).get("market_cap"),
        "pe_ratio": (valuation_snapshot or {}).get("pe_ratio"),
        "forward_pe": (valuation_snapshot or {}).get("forward_pe"),
        "peg_ratio": (valuation_snapshot or {}).get("peg_ratio"),
        "dividend_yield": (valuation_snapshot or {}).get("dividend_yield"),
        "revenue_growth_yoy": (latest_ratios or {}).get("revenue_growth_yoy"),
        "net_margin": (latest_ratios or {}).get("net_margin"),
        "roe": (latest_ratios or {}).get("roe"),
        "debt_to_equity": (latest_ratios or {}).get("debt_to_equity"),
        "valuation_rating": valuation_rating,
        "latest_recommendation": recommendation,
        "last_updated": today_str(),
    }

    row = _row_from_dict(row_data, DASHBOARD_HEADERS)
    row_number = _find_row_number(ws, ticker)
    if row_number:
        _write_row_at(ws, row_number, row, DASHBOARD_HEADERS)
    else:
        _append_row(ws, row, DASHBOARD_HEADERS)
    logger.info("Dashboard: upserted %s", ticker)


# --------------------------------------------------------------------------
# Watchlist (auto columns upserted; manual columns never touched)
# --------------------------------------------------------------------------

def ensure_watchlist_row(ticker: str, company_name: Optional[str] = None, sector: Optional[str] = None) -> None:
    """Create a Watchlist row for `ticker` if one doesn't exist yet, with
    manual columns left blank for the user to fill in themselves. If the
    row already exists, only the auto columns (name/sector) are
    refreshed -- called from the automated sync path (sync_ticker_full),
    so the manual columns are never read, written, or otherwise touched
    here. Deliberate user edits go through update_watchlist_manual_fields
    instead, not this function.
    """
    ticker = ticker.upper()
    ws = _get_or_create_worksheet(WATCHLIST_TAB, WATCHLIST_HEADERS)
    row_number = _find_row_number(ws, ticker)

    if row_number:
        _write_row_at(ws, row_number, [ticker, company_name, sector], WATCHLIST_HEADERS)
        return

    new_row = [ticker, company_name, sector] + [None] * len(WATCHLIST_MANUAL_HEADERS)
    _append_row(ws, new_row, WATCHLIST_HEADERS)
    logger.info("Watchlist: created row for %s", ticker)


def get_watchlist_record(ticker: str) -> Optional[dict]:
    """Current Watchlist row for `ticker` (auto + manual columns), or
    None. Used by the dashboard app to pre-fill the manual-field edit
    form with whatever's already saved."""
    ticker = ticker.upper()
    ws = _get_or_create_worksheet(WATCHLIST_TAB, WATCHLIST_HEADERS)
    for record in _get_all_records(ws, WATCHLIST_HEADERS):
        if record.get("ticker") == ticker:
            return record
    return None


def update_watchlist_manual_fields(ticker: str, **fields) -> None:
    """Write user-supplied values for the Watchlist's manual columns
    (investment_thesis, catalysts, risks, target_price, personal_rating,
    notes) and save immediately.

    This is the one place those columns are written from code -- and
    only in direct response to a deliberate edit in the dashboard app,
    never from the automated daily/quarterly sync (see
    ensure_watchlist_row's docstring).
    """
    ticker = ticker.upper()
    ws = _get_or_create_worksheet(WATCHLIST_TAB, WATCHLIST_HEADERS)
    row_number = _find_row_number(ws, ticker)
    if row_number is None:
        raise ValueError(f"{ticker} has no Watchlist row yet -- call ensure_watchlist_row first")

    unknown = set(fields) - set(WATCHLIST_MANUAL_HEADERS)
    if unknown:
        raise ValueError(f"Unknown Watchlist field(s): {unknown}")

    for header, value in fields.items():
        col_idx = WATCHLIST_HEADERS.index(header) + 1
        ws.cell(row=row_number, column=col_idx, value=_to_cell(value, header))
    save_workbook()
    logger.info("Watchlist: updated manual fields for %s (%s)", ticker, ", ".join(fields))


# --------------------------------------------------------------------------
# Full ticker sync -- the one entry point add_company.py and main.py both
# call, so onboarding a new ticker and the recurring daily/quarterly
# automation can't drift into different sync behavior.
# --------------------------------------------------------------------------

def sync_ticker_full(ticker: str, run_financials: bool = True, run_ai_analysis: bool = True) -> dict:
    """Sync every tab for one ticker in a single call, then save the
    workbook to disk.

    Financials, valuation, and news are each fetched exactly once here
    and reused across every tab and the AI analysis that needs them --
    without this, financials would be fetched 3x and news would be
    re-classified by the AI provider 3x per ticker per run (once each
    for their own tab and the AI narrative).

    run_financials=False skips writing to the Financials/Ratios tabs
    (main.py's daily runs use this) -- new quarters only appear a few
    times a year, so checking for them daily across every tracked
    ticker is pointless work; the spec's own automation section only
    calls for financials/ratios on the quarterly cadence.

    Saves to disk at the end of every ticker (not just at the end of
    the whole run) so a crash partway through a multi-ticker run
    doesn't lose progress already made -- the local-file equivalent of
    how each individual write used to be immediately durable via the
    Sheets API.

    Returns a summary dict (counts of what changed) for the caller to
    log or report to the user.
    """
    ticker = ticker.upper()
    summary = {"ticker": ticker}

    financials_df = financials.get_quarterly_financials(ticker)
    news_df = news.get_news(ticker, enrich=True)
    profile = valuation.get_company_profile(ticker)
    val_snapshot = valuation.get_valuation_snapshot(ticker)

    if run_financials:
        summary["new_financial_quarters"] = sync_financials(ticker, financials_df=financials_df)
        summary["new_ratio_quarters"] = sync_ratios(ticker, financials_df=financials_df)
    else:
        summary["new_financial_quarters"] = 0
        summary["new_ratio_quarters"] = 0
    summary["new_news_items"] = sync_news(ticker, news_df=news_df)

    upsert_valuation(ticker, val_snapshot)

    record = tickers.get_ticker_record(ticker)
    company_name = (record or {}).get("name") or profile.get("name") or ticker
    sector = (record or {}).get("sector") or profile.get("sector")
    ensure_watchlist_row(ticker, company_name=company_name, sector=sector)

    latest_analysis = get_latest_narrative(ticker)
    if run_ai_analysis:
        evidence = analysis.build_evidence(ticker, financials_df=financials_df, news_df=news_df)
        fresh_analysis = analysis.generate_analysis(ticker, previous_analysis=latest_analysis, evidence=evidence)
        if fresh_analysis:
            append_narrative(ticker, fresh_analysis)
            latest_analysis = fresh_analysis
        summary["ai_analysis_generated"] = fresh_analysis is not None
    else:
        summary["ai_analysis_generated"] = False

    latest_ratios = ratios.compute_ratios(financials_df)
    latest_ratios = latest_ratios.iloc[-1].to_dict() if not latest_ratios.empty else {}

    upsert_dashboard(ticker, profile, val_snapshot, latest_ratios, latest_analysis)

    save_workbook()

    summary["success"] = True
    return summary
